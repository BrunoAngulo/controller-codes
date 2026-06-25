import os
from pathlib import Path
import subprocess
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".tsv", ".txt"}

OUTPUT_COLUMNS = [
    "colegio",
    "grado",
    "grupo",
    "alumno_id",
    "nombre",
    "apellido_paterno",
    "apellido_materno",
    "codigo_acceso",
    "login",
]

OLE_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
ZIP_XLSX_MAGIC = b"PK"


def normalize_column_name(column: object) -> str:
    return str(column).strip().lower()


def read_file_start(input_path: Path) -> bytes:
    return input_path.read_bytes()[:2048].lstrip()


def read_legacy_xls(input_path: Path) -> pd.DataFrame:
    try:
        return pd.read_excel(input_path, engine="xlrd")
    except ImportError as exc:
        raise ValueError(
            "Para convertir archivos Excel 97-2003 (.xls), instala xlrd con: "
            "python -m pip install --user xlrd"
        ) from exc


def read_text_table(input_path: Path) -> pd.DataFrame:
    first_bytes = read_file_start(input_path).lower()

    if first_bytes.startswith(b"<"):
        try:
            tables = pd.read_html(input_path)
            if tables:
                return tables[0]
        except Exception:
            pass

    for encoding in ("utf-8-sig", "latin1"):
        try:
            return pd.read_csv(
                input_path,
                sep=None,
                engine="python",
                encoding=encoding,
            )
        except UnicodeDecodeError:
            continue

    return pd.read_csv(input_path, sep=None, engine="python", encoding="latin1")


def read_input_file(input_path: Path) -> pd.DataFrame:
    suffix = input_path.suffix.lower()
    first_bytes = read_file_start(input_path)

    if suffix in {".xlsx", ".xlsm"} or first_bytes.startswith(ZIP_XLSX_MAGIC):
        try:
            return pd.read_excel(input_path, engine="openpyxl")
        except Exception:
            return read_text_table(input_path)

    if suffix == ".xls":
        if first_bytes.startswith(OLE_XLS_MAGIC):
            return read_legacy_xls(input_path)

        try:
            return pd.read_excel(input_path, engine="xlrd")
        except Exception:
            return read_text_table(input_path)

    if suffix in {".csv", ".tsv", ".txt"}:
        return read_text_table(input_path)

    raise ValueError(
        "Formato no soportado. Usa .xlsx, .xlsm, .xls, .csv, .tsv o .txt."
    )


def transform_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [normalize_column_name(column) for column in df.columns]

    missing_columns = get_missing_columns(df)
    if missing_columns:
        available_columns = ", ".join(df.columns)
        missing_columns_text = ", ".join(missing_columns)
        raise ValueError(
            "Faltan columnas requeridas: "
            f"{missing_columns_text}. Columnas disponibles: {available_columns}"
        )

    return df[OUTPUT_COLUMNS]


def get_missing_columns(df: pd.DataFrame) -> list[str]:
    normalized_columns = {
        normalize_column_name(column)
        for column in df.columns
    }
    return [
        column
        for column in OUTPUT_COLUMNS
        if column not in normalized_columns
    ]


def format_workbook(output_path: Path) -> None:
    wb = load_workbook(output_path)
    ws = wb.active

    thin = Side(style="thin", color="D9E2EC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    body_alignment = Alignment(vertical="center")
    header_alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = body_alignment

            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)

    wb.save(output_path)


def build_formatted_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_formateado.xlsx")


def build_legacy_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_2003.xls")


def convert_xlsx_to_xls(input_path: Path, output_path: Path) -> None:
    powershell_script = r"""
$ErrorActionPreference = "Stop"
$inputPath = [System.IO.Path]::GetFullPath($env:FORMAT_INPUT_PATH)
$outputPath = [System.IO.Path]::GetFullPath($env:FORMAT_OUTPUT_PATH)
$excel = $null
$workbook = $null

try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($inputPath, 0, $true)
    $workbook.SaveAs($outputPath, 56)
    $workbook.Close($false)
    $workbook = $null
}
finally {
    if ($null -ne $workbook) {
        $workbook.Close($false)
        [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($workbook)
    }

    if ($null -ne $excel) {
        $excel.Quit()
        [void][Runtime.InteropServices.Marshal]::FinalReleaseComObject($excel)
    }

    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}
"""
    env = os.environ.copy()
    env["FORMAT_INPUT_PATH"] = str(input_path.resolve())
    env["FORMAT_OUTPUT_PATH"] = str(output_path.resolve())

    if output_path.exists():
        output_path.unlink()

    result = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            powershell_script,
        ],
        capture_output=True,
        text=True,
        env=env,
        check=False,
    )

    if result.returncode != 0 or not output_path.exists():
        error_text = result.stderr.strip() or result.stdout.strip()
        raise ValueError(
            "No se pudo convertir a Excel 97-2003 con Microsoft Excel."
            + (f" Detalle: {error_text}" if error_text else "")
        )


def collect_input_files(paths: list[str]) -> list[Path]:
    input_files = []

    for raw_path in paths:
        path = Path(raw_path).resolve()

        if not path.exists():
            print(f"No se encontro: {path}")
            continue

        if path.is_dir():
            folder_files = [
                file
                for file in path.iterdir()
                if file.is_file()
                and file.suffix.lower() in SUPPORTED_EXTENSIONS
                and not file.stem.endswith(("_formateado", "_2003"))
            ]
            input_files.extend(sorted(folder_files))
            continue

        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            print(f"Formato no soportado, se omite: {path}")
            continue

        input_files.append(path)

    return input_files


def process_file(input_path: Path) -> bool:
    try:
        df = read_input_file(input_path)

        if not get_missing_columns(df):
            output_path = build_formatted_output_path(input_path)
            formatted_df = transform_columns(df)
            formatted_df.to_excel(output_path, index=False)
            format_workbook(output_path)
            print(f"Archivo con estructura detectada y formateado: {output_path}")
            return True

        if input_path.suffix.lower() == ".xlsx":
            output_path = build_legacy_output_path(input_path)
            convert_xlsx_to_xls(input_path, output_path)
            print(f"Archivo sin la estructura requerida convertido a Excel 97-2003: {output_path}")
            return True

        missing_columns = ", ".join(get_missing_columns(df))
        raise ValueError(
            f"Faltan columnas requeridas: {missing_columns}. "
            "Solo los archivos .xlsx sin esa estructura se convierten a Excel 97-2003."
        )
    except Exception as exc:
        print(f"Error: {exc}")
        return False


def main() -> int:
    if len(sys.argv) < 2:
        print("Uso: python format_report.py <archivo_o_carpeta> [mas_archivos...]")
        return 1

    input_files = collect_input_files(sys.argv[1:])

    if not input_files:
        print("No se encontraron archivos para formatear.")
        return 1

    print(f"Archivos a procesar: {len(input_files)}")

    has_error = False
    for input_path in input_files:
        print("")
        print(f"Procesando: {input_path}")
        if not process_file(input_path):
            has_error = True

    if has_error:
        print("")
        print("Proceso terminado con errores en uno o mas archivos.")
        return 1

    print("")
    print("Proceso terminado correctamente.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
